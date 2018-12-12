from openpyxl import load_workbook


class XLSXParser:
    def __init__(self, file_path: str):
        self.file_path = file_path

    def parse_headline(self):
        wb = load_workbook(self.file_path)
        sheets = []
        col = {}
        for sheet_name in wb.sheetnames:
            sheets.append(sheet_name)
            ws = wb.get_sheet_by_name(sheet_name)
            for row in ws.iter_rows():
                res = []
                for cell in row:
                    res.append(str(cell.value))
                col[sheet_name] = res
                break

        return {'sheets': sheets, 'col': col}

    def get_data_by_sheet_name(self, sheet_name):
        wb = load_workbook(self.file_path)
        ws = wb.get_sheet_by_name(sheet_name)
        res = []
        for i, row in enumerate(ws.iter_rows()):
            if i == 0:
                continue
            tmp_row = []
            for cell in row:
                tmp_row.append(str(cell.value))
            res.append(tmp_row)
        return res
